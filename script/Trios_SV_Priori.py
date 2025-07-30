#!/usr/bin/env python3

# usage: python3 Trios_SV_Priori.py --annotSR annotsvfile --output SRR645629_SV.xlsx
import pandas as pd
import argparse
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

def standardize_zygocity(value):
    """Convert various genotype formats to standardized Zygocity values"""
    if pd.isna(value):
        return "Nocall"    
    value = str(value).strip() 
    # Handle common cases first
    if value == '0/0':
        return "Hom_ref"
    elif value == '0/1':
        return "Het"
    elif value == '1/1':
        return "Hom"
    elif value == './.':
        return "Nocall"
    
    # Handle less common cases
    elif value in ['0|1', '1|0']:
        return "Het_phased"
    elif value == '1/2':
        return "Het_alt"
    
    # Fallback for unexpected values
    return value
def assign_inheritance(gt_f, gt_m, gt_c, chr):
    """Determine inheritance pattern based on genotypes"""
    if pd.isna(gt_f) or pd.isna(gt_m) or pd.isna(gt_c):
        return "Unknown"
    
    gt_f = str(gt_f).split(":")[0]
    gt_m = str(gt_m).split(":")[0]
    gt_c = str(gt_c).split(":")[0]

    # De Novo
    if (gt_c in ["0/1", "1/1"]) and (gt_f == "0/0" and gt_m == "0/0"):
        return "DeNovo"
    
    # Autosomal Dominant
    if gt_c in ["0/1", "1/1"] and (gt_f != "0/0" or gt_m != "0/0"):
        return "AD"
    
    # Autosomal Recessive
    if gt_c == "1/1" and gt_f == "0/1" and gt_m == "0/1":
        return "AR"
    
    # UPD
    if gt_c == "1/1" and ((gt_f == "1/1" and gt_m == "0/0") or (gt_f == "0/0" and gt_m == "1/1")):
        return "UPD"
    
    # X/Y-linked
    if chr == "X":
        if gt_c == "1" and gt_m == "0/1":
            return "XL"
    elif chr == "Y":
        if gt_f == "1" and gt_c == "1":
            return "YL"
    
    return "Unknown"

def format_ranked_sv(df):
    """Format the Ranked_SV dataframe according to specifications"""
    # Define required columns for SV_Position calculation
    required_columns = ['SV_type', 'SV_chrom', 'SV_start', 'SV_end', 'SV_length']
    # Create empty DataFrame with all required columns if input is empty
    if df.empty:
        all_columns = required_columns + [
            'Gene_name', 'Gene_count', 'CytoBand', 'Zygocity', 'P_gain_phen',
            'P_gain_source', 'P_loss_phen', 'P_loss_source', 'HI', 'TS',
            'DDD_HI_percent', 'OMIM_ID', 'GnomAD_pLI', 'ACMG_classification'
        ]
        df = pd.DataFrame(columns=all_columns)
    # Ensure SV_length is numeric
    if 'SV_length' in df.columns:
        df['SV_length'] = pd.to_numeric(df['SV_length'], errors='coerce')    
    # Only proceed with SV_Position calculation if required columns exist
    if all(col in df.columns for col in required_columns):
        # Create SV_Position column
        df['SV_Position'] = df['SV_type'].astype(str) + '_' + df['SV_chrom'].astype(str) + ':' + df['SV_start'].astype(str) + '-' + df['SV_end'].astype(str)
        # Format SV_length - handle both positive and negative values appropriately
        def format_sv_length(length):
            try:
                length = float(length)
                abs_length = abs(length)
                abs_length_kb = abs_length / 1000
                # Determine the appropriate unit and scale
                if abs_length_kb >= 1000:  # If ≥1000kb (1MB)
                    scaled_length = abs_length_kb / 1000
                    unit = "Mb"
                else:
                    scaled_length = abs_length_kb
                    unit = "kb"
                # Round to 1 decimal place and format with sign
                formatted_length = round(scaled_length, 1)
                if length < 0:
                    return f"-{formatted_length}{unit}"
                else:
                    return f"{formatted_length}{unit}"
            except (ValueError, TypeError):
                return str(length)  # Return original if can't convert to number
        df['SV_length'] = df['SV_length'].apply(format_sv_length)
    else:
        # If required columns are missing, create empty formatted columns
        df['SV_Position'] = ''
        df['SV_length'] = ''

    # Rename columns
    rename_map = {
        'P_gain_phen': 'Pathogenic_Gain_Disease',
        'P_gain_source': 'Pathogenic_Gain_Source',
        'P_loss_phen': 'Pathogenic_Loss_Disease',
        'P_loss_source': 'Pathogenic_Loss_Source',
        'HI': 'ClinGene_HI_score',
        'TS': 'ClinGene_TS_score',
        'DDD_HI_percent': 'DECIPHER_HI_score',
        'GnomAD_pLI': 'pLI'
    }
    # Only rename columns that exist in the dataframe
    rename_map = {k: v for k, v in rename_map.items() if k in df.columns}
    df = df.rename(columns=rename_map)
    # Select and reorder columns
    output_columns = [
        'SV_Position', 'SV_length', 'Gene_name', 'Gene_count', 'CytoBand', 'Zygocity',
        'INHERITANCE', 'Pathogenic_Gain_Disease', 'Pathogenic_Gain_Source', 'Pathogenic_Loss_Disease',
        'Pathogenic_Loss_Source', 'OMIM_ID', 'ClinGene_HI_score', 'ClinGene_TS_score',
        'DECIPHER_HI_score', 'pLI', 'ACMG_classification'
    ]
    
    # Only keep columns that exist in the dataframe
    existing_columns = [col for col in output_columns if col in df.columns]
    return df[existing_columns]

def apply_tab_color(workbook, sheet_name, color):
    """Apply color to the worksheet tab"""
    if sheet_name in workbook.sheetnames:
        workbook[sheet_name].sheet_properties.tabColor = color

def auto_adjust_columns(worksheet):
    """Auto-adjust column widths"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Merge and process genomic files')
    parser.add_argument('--annotSR', required=True, help='Annotated input file (TSV)')
    parser.add_argument('--output', required=True, help='Output XLSX file path')
    args = parser.parse_args()

    # Read input files
    df = pd.read_csv(args.annotSR, sep='\t', low_memory=False)
    # Determine inheritance from columns 14,15,16 (0-based index 13,14,15)
    if len(df.columns) >= 16:
        print("Determining inheritance patterns from genotype columns...")
        # Get chromosome column (assuming it's column 0)
        chrom_col = df.columns[1]
        son_col = df.columns[14]
        mother_col = df.columns[15]
        father_col = df.columns[16]
        
        # Apply inheritance determination
        df['INHERITANCE'] = df.apply(
            lambda row: assign_inheritance(
                row[father_col], 
                row[mother_col], 
                row[son_col], 
                row[chrom_col]
            ),
            axis=1
        )
    else:
        print("Warning: Input file doesn't have enough columns for inheritance determination")
        df['INHERITANCE'] = 'Unknown'

    # # Process INFO column to extract specific fields
    # print("Processing INFO column...")
    # def extract_info_fields(info):
    #     fields = {}
    #     for part in info.split(';'):
    #         if '=' in part:
    #             key, value = part.split('=', 1)
    #             if key in ['INHERITANCE']:
    #                 fields[key] = value
    #     return pd.Series(fields)

    # # Extract fields and add as new columns
    # info_fields = df['INFO'].apply(extract_info_fields)
    # df = pd.concat([df, info_fields], axis=1)
   
    print("Standardizing Zygocity column...")
    zyg_col = df.columns[14]
    df['Zygocity'] = df[zyg_col].str.split(':').str[0].apply(standardize_zygocity)

    # Split the column into genotype and copy number
    #zyg_split = df[zyg_col].str.split(':', expand=True)
    #df['Zygocity'] = zyg_split[0].apply(standardize_zygocity)  # Standardize the genotype
    #df['Copy_Number'] = pd.to_numeric(zyg_split[1], errors='coerce')  # Extract copy number after colon

    # Handle cases where there was no colon (set Copy_Number to NaN)
    #df.loc[~df[zyg_col].str.contains(':'), 'Copy_Number'] = pd.NA

    # Remove unwanted columns
    columns_to_drop = [
        'ID', 'QUAL', 'FILTER', 'FORMAT', 'Closest_left', 
        'Closest_right', 'Tx', 'Tx_version', 'Tx_start', 'Tx_end', 'Exon_count', 
        'Location', 'Location2', 'Dist_nearest_SS', 'Nearest_SS_type', 'Intersect_start', 
        'Intersect_end', 'RE_gene', 'P_ins_phen', 'P_ins_hpo', 'P_ins_source', 
        'P_ins_coord', 'B_loss_AFmax', 'B_ins_source', 'B_ins_coord', 'B_ins_AFmax', 
        'B_inv_source', 'B_inv_coord', 'B_inv_AFmax', 'po_B_gain_allG_source', 
        'po_B_gain_allG_coord', 'po_B_gain_someG_source', 'po_B_gain_someG_coord', 
        'po_B_loss_allG_source', 'po_B_loss_allG_coord', 'po_B_loss_someG_source', 
        'po_B_loss_someG_coord', 'GC_content_left', 'GC_content_right', 
        'Repeat_coord_left', 'Repeat_type_left', 'Repeat_coord_right', 
        'Repeat_type_right', 'Gap_left', 'Gap_right', 'SegDup_left', 'SegDup_right', 
        'ENCODE_blacklist_left', 'ENCODE_blacklist_characteristics_left', 
        'ENCODE_blacklist_right', 'ENCODE_blacklist_characteristics_right', 
        'B_gain_AFmax', 'ACMG', 'ExAC_synZ', 'ExAC_misZ', 'NCBI_gene_ID'
    ]

    # Only drop columns that exist in the dataframe
    columns_to_drop = [col for col in columns_to_drop if col in df.columns]
    df = df.drop(columns=columns_to_drop)

    # Add ACMG classification tags
    print("Adding ACMG classification...")
    acmg_mapping = {
        '1': 'Benign',
        '2': 'Likely Benign',
        '3': 'VUS',
        '4': 'Likely Pathogenic',
        '5': 'Pathogenic'
    }

    # Convert HI/TS to numeric if columns exist
    for col in ['HI', 'TS']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # Create new classification column if ACMG_class exists
    if 'ACMG_class' in df.columns:
        # Convert ACMG_class to string and clean it
        df['ACMG_class'] = df['ACMG_class'].astype(str).str.strip()
        df['ACMG_classification'] = df['ACMG_class'].map(acmg_mapping)
        # For any unmapped values, keep original
        df['ACMG_classification'] = df['ACMG_classification'].fillna(df['ACMG_class'])
    
    # Create ACMG filtered sheet - only numeric 3,4,5 values
    print("Creating ACMG filtered sheet...")
    acmg_filtered = pd.DataFrame()

    if 'ACMG_class' in df.columns:
        # Convert ACMG_class to string and strip whitespace for comparison
        df['ACMG_class'] = df['ACMG_class'].astype(str).str.strip()
        # Filter for numeric 3,4,5 only (not "full=3", etc.)
        acmg_numeric = df['ACMG_class'].isin(['3', '4', '5'])
        # Combined filter for numeric ACMG 3,4,5
        acmg_conditions = (
            acmg_numeric
        )
        acmg_filtered = df[acmg_conditions]
    # Create pathogenic_SV sheet if P_gain_coord or P_loss_coord has data
    pathogenic_sv = None
    if not acmg_filtered.empty:
        coord_columns = ['P_gain_coord', 'P_loss_coord', 'po_P_gain_coord', 'po_P_loss_coord']
        existing_coord_cols = [col for col in coord_columns if col in acmg_filtered.columns]
        
        if existing_coord_cols:
            print("Creating pathogenic_SV sheet...")
            coord_condition = acmg_filtered[existing_coord_cols].notna().any(axis=1)
            pathogenic_sv = acmg_filtered[coord_condition]

   # Create OMIM_SV sheet if OMIM_morbid or OMIM_morbid_candidate contains "YES"
    omim_sv = None
    if pathogenic_sv is not None and not pathogenic_sv.empty:
        omim_columns = ['OMIM_morbid', 'OMIM_morbid_candidate']
        existing_omim_cols = [col for col in omim_columns if col in pathogenic_sv.columns]
        
        if existing_omim_cols:
            print("Creating OMIM_SV sheet...")
            # Convert columns to string and handle NaN values
            omim_df = pathogenic_sv[existing_omim_cols].astype(str)
            omim_condition = omim_df.apply(
                lambda x: x.str.upper().str.strip() == 'YES'
            ).any(axis=1)
            omim_sv = pathogenic_sv[omim_condition]

    # Create Ranked_SV sheet for HI/TS scores 2 or 3 from OMIM_SV (numeric comparison)
    ranked_sv = None
    if omim_sv is not None and not omim_sv.empty:
        hi_ts_columns = ['HI', 'TS']
        existing_hi_ts_cols = [col for col in hi_ts_columns if col in omim_sv.columns]
        
        if existing_hi_ts_cols:
            print("Creating Ranked_SV sheet from OMIM_SV...")
            # Numeric comparison for HI/TS scores
            hi_ts_condition = omim_sv[existing_hi_ts_cols].apply(
                lambda x: x.isin([2, 3])
            ).any(axis=1)
            ranked_sv = omim_sv[hi_ts_condition]
            
            # Filter and sort by Gene_count if column exists
            if 'Gene_count' in ranked_sv.columns:
                # Remove rows with Gene_count < 10
                ranked_sv = ranked_sv[ranked_sv['Gene_count'] > 10]
                print(f"Filtered to {len(ranked_sv)} rows with Gene_count > 10")
                
                # Sort by Gene_count in descending order
                #ranked_cnv = ranked_cnv.sort_values('Gene_count', ascending=False)
                #print(f"Sorted Ranked_CNV by Gene_count (high to low) {len(ranked_cnv)}")

                inheritance_order = ['DeNovo', 'XL', 'AD', 'AR', 'XY', 'UPD', 'Unknown']
                # Convert INHERITANCE to categorical with specified order
                if 'INHERITANCE' in ranked_sv.columns:
                    ranked_sv['INHERITANCE'] = pd.Categorical(
                        ranked_sv['INHERITANCE'],
                        categories=inheritance_order,
                        ordered=True
                    )
    
                    # Sort by INHERITANCE (and maintain other sorting if needed)
                    ranked_sv = ranked_sv.sort_values(by=['INHERITANCE', 'Gene_count'], ascending=[True, True])
                    print(f"Sorted Ranked_SV by Gene_count (high to low) {len(ranked_sv)}")
                else:
                    print("INHERITANCE column not found - skipping inheritance-based sorting")
    
    # Save output to Excel with sheets
    print(f"Writing output to {args.output}...")
    with pd.ExcelWriter(args.output, engine='openpyxl') as writer:
        # Write all sheets first
        df.to_excel(writer, sheet_name='All_Data', index=False)
        
        if not acmg_filtered.empty:
            acmg_filtered.to_excel(writer, sheet_name='ACMG', index=False)
        
        if pathogenic_sv is not None and not pathogenic_sv.empty:
            pathogenic_sv.to_excel(writer, sheet_name='pathogenic_SV', index=False)
            print(f"Added pathogenic_SV sheet with {len(pathogenic_sv)} rows")
        
        if omim_sv is not None and not omim_sv.empty:
            omim_sv.to_excel(writer, sheet_name='OMIM_SV', index=False)
            print(f"Added OMIM_SV sheet with {len(omim_sv)} rows")

        # Always create Ranked_SV sheet, even if empty    
        ranked_sv_to_write = ranked_sv if (ranked_sv is not None and not ranked_sv.empty) else pd.DataFrame()
        ranked_sv_formatted = format_ranked_sv(ranked_sv_to_write)
        # Create the sheet with formatted columns
        ranked_sv_formatted.to_excel(writer, sheet_name='Ranked_SV', index=False)
        workbook = writer.book
        worksheet = workbook['Ranked_SV'] 
        # Apply green color only to the tab
        apply_tab_color(workbook, 'Ranked_SV', '92D050')  # Green color
            
        # Auto-adjust column widths
        auto_adjust_columns(worksheet)
            
        # Set Ranked_SV as the active sheet
        workbook.active = workbook['Ranked_SV']
        if not ranked_sv_formatted.empty:
            print(f"Added formatted Ranked_SV sheet with {len(ranked_sv_formatted)} rows (green tab)")
          
        else:
            print("Created empty Ranked_SV sheet with proper formatting (no variants met criteria)")
    
    print(f"Output saved with {len(df)} rows in All_Data")
    print("Done!")

if __name__ == '__main__':
    main()
